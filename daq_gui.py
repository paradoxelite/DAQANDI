import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import threading
import time
from daq_interface import DAQInterface
import math
import numpy as np
import os
from datetime import datetime
from rt_daq_interface import RTDAQInterface
import queue
import sys
from openpyxl import Workbook
from PyQt5.QtWidgets import QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, QCheckBox, QLabel
from PyQt5.QtCore import Qt
import pyqtgraph as pg

class DAQGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("DAQ Interface")
        self.root.geometry("1000x700")
        
        # Configurar el protocolo de cierre antes de cualquier otra operación
        self.root.protocol("WM_DELETE_WINDOW", self.on_closing)
        
        # Variables para control de tiempo y actualización
        self.last_update_time = 0
        self.update_interval = 0.1  # 100ms entre actualizaciones de visualización
        self.read_sampling_period = 0.0002  # 200 microsegundos
        self.write_sampling_period = 0.00001  # 10 microsegundos
        self.max_points = 1000  # Número máximo de puntos en las gráficas
        
        # Arrays para almacenar datos
        self.time_data = np.zeros(self.max_points)
        self.digital_data = np.zeros(self.max_points)
        self.analog_data = np.zeros(self.max_points)
        self.command_data = {
            'Y1': np.zeros(self.max_points),
            'Y2': np.zeros(self.max_points),
            'Y3': np.zeros(self.max_points),
            'Y4': np.zeros(self.max_points),
            'Y5': np.zeros(self.max_points)
        }
        
        # Buffer para guardar datos
        self.data_buffer = {
            'time': [],
            'digital_value': [],
            'analog_value': [],
            'Y1': [],
            'Y2': [],
            'Y3': [],
            'Y4': [],
            'Y5': []
        }
        
        # Configurar el icono de la aplicación
        try:
            current_dir = os.path.dirname(os.path.abspath(__file__))
            icon_path = os.path.join(current_dir, 'icon.ico')
            self.root.iconbitmap(icon_path)
        except Exception as e:
            print(f"Error al cargar el icono: {e}")
        
        # Configurar estilo moderno
        style = ttk.Style()
        
        # Definir colores modernos
        self.colors = {
            'primary': '#1976D2',      # Azul principal más oscuro
            'secondary': '#D32F2F',    # Rojo secundario más oscuro
            'success': '#2E7D32',      # Verde más oscuro
            'warning': '#F57F17',      # Amarillo más oscuro
            'error': '#C62828',        # Rojo más oscuro
            'background': '#FAFAFA',   # Fondo claro
            'surface': '#FFFFFF',      # Superficie blanca
            'text': '#212121',         # Texto principal
            'text_secondary': '#757575', # Texto secundario
            'button_text': '#FFFFFF',   # Texto de botones
            'button_bg': '#1976D2',     # Fondo de botones
            'button_bg_secondary': '#D32F2F'  # Fondo de botones secundarios
        }
        
        # Configurar estilos de widgets
        style.configure('TFrame', background=self.colors['background'])
        style.configure('TLabel', 
                       background=self.colors['background'],
                       foreground=self.colors['text'],
                       font=('Segoe UI', 10))
        
        # Estilos personalizados para botones
        style.configure('Accent.TButton',
                       background=self.colors['button_bg'],
                       foreground=self.colors['button_text'],
                       font=('Segoe UI', 10, 'bold'),
                       padding=10)
        
        style.configure('Secondary.TButton',
                       background=self.colors['button_bg_secondary'],
                       foreground=self.colors['button_text'],
                       font=('Segoe UI', 10, 'bold'),
                       padding=10)
        
        style.configure('TButton',
                       background=self.colors['button_bg'],
                       foreground=self.colors['button_text'],
                       font=('Segoe UI', 10, 'bold'),
                       padding=10)
        
        style.configure('TEntry',
                       fieldbackground=self.colors['surface'],
                       foreground=self.colors['text'],
                       font=('Segoe UI', 10))
        
        style.configure('TLabelframe',
                       background=self.colors['background'],
                       foreground=self.colors['text'],
                       font=('Segoe UI', 10))
        
        style.configure('TLabelframe.Label',
                       background=self.colors['background'],
                       foreground=self.colors['text'],
                       font=('Segoe UI', 10, 'bold'))
        
        style.configure('TNotebook',
                       background=self.colors['background'])
        
        style.configure('TNotebook.Tab',
                       background=self.colors['surface'],
                       foreground=self.colors['text'],
                       font=('Segoe UI', 10),
                       padding=[10, 5])
        
        # Inicializar DAQ con interfaz en tiempo real
        self.daq = RTDAQInterface()
        if not self.daq.connect_device():
            messagebox.showerror("Error", "Failed to connect to device")
            self.on_closing()
            return
            
        # Variables para control
        self.digital_value = tk.IntVar(value=0)
        self.analog_channel = tk.IntVar(value=1)
        self.analog_value = tk.DoubleVar(value=0.0)
        self.wave_type = tk.StringVar(value="sine")
        self.frequency = tk.DoubleVar(value=1.0)
        self.amplitude = tk.DoubleVar(value=2.5)
        self.offset = tk.DoubleVar(value=0.0)
        
        # Variables para lectura
        self.reading_active = False
        self.read_thread = None
        self.start_time = None
        self.measurement_started = False
        
        # Variables para comandos
        self.commands_active = False
        self.command_thread = None
        self.command_start_time = None
        self.command_expressions = {
            'Y1': tk.StringVar(value="0"),
            'Y2': tk.StringVar(value="0"),
            'Y3': tk.StringVar(value="0"),
            'Y4': tk.StringVar(value="0"),
            'Y5': tk.StringVar(value="0")
        }
        
        # Variables para almacenar valores de entrada
        self.input_values = {
            'X1': 0.0,  # AN1
            'X2': 0.0,  # AN2
            'X3': 0.0,  # AN3
            'X4': 0.0,  # AN4
            'X5': 0,    # Digital
            't': 0.0,   # Tiempo
            'sin': math.sin,
            'cos': math.cos,
            'tan': math.tan,
            'sqrt': math.sqrt,
            'abs': abs,
            'pi': math.pi,
            'e': math.e
        }
        
        # Inicializar ventana de gráficas
        self.qt_app = QApplication(sys.argv)
        self.graphs_window = GraphsWindow()
        self.graphs_window.show()
        
        # Vincular movimiento de ventana principal con ventana de gráficas
        self.root.bind('<Configure>', self.on_window_move)
        
        self.create_widgets()
        
        # Reiniciar buffers y mostrar display digital al inicio
        self.reset_digital()
        self.update_digital_display(0)

    def create_widgets(self):
        # Crear notebook para pestañas
        self.notebook = ttk.Notebook(self.root)
        self.notebook.pack(expand=True, fill='both', padx=5, pady=5)
        
        # Pestaña de salidas digitales
        digital_frame = ttk.Frame(self.notebook)
        self.notebook.add(digital_frame, text="Digital Outputs")
        self.create_digital_outputs(digital_frame)
        
        # Pestaña de salidas analógicas
        analog_frame = ttk.Frame(self.notebook)
        self.notebook.add(analog_frame, text="Analog Outputs")
        self.create_analog_outputs(analog_frame)
        
        # Pestaña de entradas
        input_frame = ttk.Frame(self.notebook)
        self.notebook.add(input_frame, text="Inputs")
        self.create_inputs(input_frame)
        
        # Pestaña de comandos
        command_frame = ttk.Frame(self.notebook)
        self.notebook.add(command_frame, text="Commands")
        self.create_commands(command_frame)
        
        # Vincular evento de cambio de pestaña
        self.notebook.bind('<<NotebookTabChanged>>', self.on_tab_changed)
        
    def create_digital_outputs(self, parent):
        # Frame principal con padding y fondo
        main_frame = ttk.Frame(parent)
        main_frame.pack(fill='both', expand=True, padx=20, pady=10)
        
        # Frame para control digital con sombra
        control_frame = ttk.LabelFrame(main_frame, text="Digital Control")
        control_frame.pack(fill='x', padx=5, pady=5)
        
        # Frame interno para mejor organización
        inner_frame = ttk.Frame(control_frame)
        inner_frame.pack(fill='x', padx=10, pady=10)
        
        # Control de valor digital con mejor espaciado y estilo
        ttk.Label(inner_frame, text="Value (0-255):").pack(side='left', padx=5)
        digital_entry = ttk.Entry(inner_frame, textvariable=self.digital_value, width=10, state='normal')
        digital_entry.pack(side='left', padx=5)
        
        # Botones con colores modernos
        write_btn = tk.Button(inner_frame, text="Write", command=self.write_digital,
                            bg=self.colors['primary'], fg='white',
                            font=('Segoe UI', 10, 'bold'),
                            relief='flat', padx=10, pady=5)
        write_btn.pack(side='left', padx=5)
        
        reset_btn = tk.Button(inner_frame, text="Reset", command=self.reset_digital,
                            bg=self.colors['secondary'], fg='white',
                            font=('Segoe UI', 10, 'bold'),
                            relief='flat', padx=10, pady=5)
        reset_btn.pack(side='left', padx=5)
        
        # Frame para visualización con fondo blanco
        display_frame = ttk.LabelFrame(main_frame, text="Digital Display")
        display_frame.pack(fill='both', expand=True, padx=5, pady=5)
        
        # Canvas para visualización binaria con fondo blanco y bordes suaves
        self.binary_canvas = tk.Canvas(display_frame, height=100, bg=self.colors['surface'], 
                                     highlightthickness=1, highlightbackground=self.colors['primary'])
        self.binary_canvas.pack(fill='x', padx=10, pady=10)
        
        # Frame para etiquetas de bits con mejor espaciado
        bits_frame = ttk.Frame(display_frame)
        bits_frame.pack(fill='x', padx=10, pady=5)
        
        # Etiquetas para los bits con colores modernos
        self.bit_labels = []
        for i in range(8):
            label = ttk.Label(bits_frame, text=f"Bit {i}: 0", 
                            foreground=self.colors['text_secondary'])
            label.pack(side='left', padx=10)
            self.bit_labels.append(label)
            
    def create_analog_outputs(self, parent):
        # Frame principal con padding
        main_frame = ttk.Frame(parent)
        main_frame.pack(fill='both', expand=True, padx=20, pady=10)
        
        # Frame para control estático con estilo moderno
        static_frame = ttk.LabelFrame(main_frame, text="Static Control")
        static_frame.pack(fill='x', padx=5, pady=5)
        
        # Frame interno para mejor organización
        inner_frame = ttk.Frame(static_frame)
        inner_frame.pack(fill='x', padx=10, pady=10)
        
        # Control de canal y valor con mejor espaciado y estilo
        ttk.Label(inner_frame, text="Channel:").pack(side='left', padx=5)
        ttk.Spinbox(inner_frame, from_=1, to=4, textvariable=self.analog_channel, 
                   width=5, state='readonly').pack(side='left', padx=5)
        
        ttk.Label(inner_frame, text="Value (-5 to 5V):").pack(side='left', padx=5)
        analog_entry = ttk.Entry(inner_frame, textvariable=self.analog_value, 
                               width=10, state='normal')
        analog_entry.pack(side='left', padx=5)
        
        # Botón con estilo moderno
        write_btn = tk.Button(inner_frame, text="Write", command=self.write_analog,
                            bg=self.colors['primary'], fg='white',
                            font=('Segoe UI', 10, 'bold'),
                            relief='flat', padx=10, pady=5)
        write_btn.pack(side='left', padx=5)
        
        # Frame para control de onda con estilo moderno
        wave_frame = ttk.LabelFrame(main_frame, text="Waveform Generator")
        wave_frame.pack(fill='x', padx=5, pady=5)
        
        # Frame para tipo de onda con mejor espaciado
        wave_type_frame = ttk.Frame(wave_frame)
        wave_type_frame.pack(fill='x', padx=10, pady=5)
        
        # Tipo de onda con mejor espaciado y estilo
        ttk.Label(wave_type_frame, text="Wave Type:").pack(side='left', padx=5)
        for wave_type in ["sine", "square", "triangle", "sawtooth"]:
            ttk.Radiobutton(wave_type_frame, text=wave_type.capitalize(), 
                          variable=self.wave_type, value=wave_type).pack(side='left', padx=5)
        
        # Frame para parámetros con mejor espaciado
        param_frame = ttk.Frame(wave_frame)
        param_frame.pack(fill='x', padx=10, pady=5)
        
        # Parámetros con mejor espaciado y estilo
        ttk.Label(param_frame, text="Frequency (Hz):").pack(side='left', padx=5)
        freq_entry = ttk.Entry(param_frame, textvariable=self.frequency, 
                             width=10, state='normal')
        freq_entry.pack(side='left', padx=5)
        
        ttk.Label(param_frame, text="Amplitude (V):").pack(side='left', padx=5)
        amp_entry = ttk.Entry(param_frame, textvariable=self.amplitude, 
                            width=10, state='normal')
        amp_entry.pack(side='left', padx=5)
        
        ttk.Label(param_frame, text="Offset (V):").pack(side='left', padx=5)
        offset_entry = ttk.Entry(param_frame, textvariable=self.offset, 
                               width=10, state='normal')
        offset_entry.pack(side='left', padx=5)
        
        # Botón con estilo moderno
        self.wave_button = tk.Button(wave_frame, text="Start Waveform", 
                            command=self.start_waveform,
                            bg=self.colors['primary'], fg='white',
                            font=('Segoe UI', 10, 'bold'),
                            relief='flat', padx=10, pady=5)
        self.wave_button.pack(pady=10)
        
        # Canvas para visualización de la onda con fondo blanco y bordes suaves
        self.wave_canvas = tk.Canvas(wave_frame, height=200, bg=self.colors['surface'],
                                   highlightthickness=1, highlightbackground=self.colors['primary'])
        self.wave_canvas.pack(fill='x', padx=10, pady=10)
        
    def create_inputs(self, parent):
        # Frame principal con scroll
        main_frame = ttk.Frame(parent)
        main_frame.pack(fill='both', expand=True)
        
        # Canvas y scrollbar para el frame principal
        canvas = tk.Canvas(main_frame, bg=self.colors['background'])
        scrollbar = ttk.Scrollbar(main_frame, orient="vertical", command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas)
        
        # Configurar el scroll
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        # Configurar el canvas para que se expanda con la ventana
        def configure_canvas(event):
            canvas.itemconfig(canvas_frame, width=event.width)
        
        canvas.bind('<Configure>', configure_canvas)
        
        canvas_frame = canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        # Pack del canvas y scrollbar
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        # Frame para entradas digitales
        digital_frame = ttk.LabelFrame(scrollable_frame, text="Digital Inputs")
        digital_frame.pack(fill='x', padx=5, pady=5)
        
        # Frame para control de medición
        control_frame = ttk.Frame(digital_frame)
        control_frame.pack(fill='x', padx=10, pady=5)
        
        # Botón de interrupción con mejor estilo
        self.interrupt_button = tk.Button(control_frame, text="Start Measurement", 
                                        command=self.toggle_measurement,
                                        bg=self.colors['primary'], fg='white',
                                        font=('Segoe UI', 10, 'bold'),
                                        relief='flat', padx=10, pady=5)
        self.interrupt_button.pack(side='left', padx=5)
        
        # Botón para guardar datos
        save_button = tk.Button(control_frame, text="Save Data", 
                              command=self.save_data_to_excel,
                              bg=self.colors['success'], fg='white',
                              font=('Segoe UI', 10, 'bold'),
                              relief='flat', padx=10, pady=5)
        save_button.pack(side='left', padx=5)
        
        # Etiqueta para mostrar valor digital
        self.digital_value_label = ttk.Label(digital_frame, text="Value: -- (0x--)")
        self.digital_value_label.pack(pady=5)
        
        # Frame para entradas analógicas
        analog_frame = ttk.LabelFrame(scrollable_frame, text="Analog Inputs")
        analog_frame.pack(fill='x', padx=5, pady=5)
        
        # Frame para controles analógicos
        control_frame = ttk.Frame(analog_frame)
        control_frame.pack(fill='x', padx=10, pady=5)
        
        # Selector de canal analógico
        ttk.Label(control_frame, text="Channel:").pack(side='left', padx=5)
        self.analog_channel_spinbox = ttk.Spinbox(control_frame, from_=1, to=4, textvariable=self.analog_channel, width=5)
        self.analog_channel_spinbox.pack(side='left', padx=5)
        self.analog_channel_spinbox.bind('<ButtonRelease-1>', self.on_analog_channel_change)
        
        # Frame para visualización del valor
        display_frame = ttk.Frame(analog_frame)
        display_frame.pack(fill='x', padx=10, pady=5)
        
        # Etiqueta para mostrar el valor actual
        ttk.Label(display_frame, text="Value:").pack(side='left', padx=5)
        self.analog_value_display = ttk.Label(display_frame, text="-- V", font=('Segoe UI', 12, 'bold'))
        self.analog_value_display.pack(side='left', padx=5)
        
    def create_commands(self, parent):
        # Frame principal con scroll
        main_frame = ttk.Frame(parent)
        main_frame.pack(fill='both', expand=True)
        
        # Canvas y scrollbar para el frame principal
        canvas = tk.Canvas(main_frame, bg=self.colors['background'])
        scrollbar = ttk.Scrollbar(main_frame, orient="vertical", command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas)
        
        # Configurar el scroll
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        # Configurar el canvas para que se expanda con la ventana
        def configure_canvas(event):
            canvas.itemconfig(canvas_frame, width=event.width)
        
        canvas.bind('<Configure>', configure_canvas)
        
        canvas_frame = canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        # Pack del canvas y scrollbar
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        # Frame para los controles de comandos
        commands_controls = ttk.Frame(scrollable_frame)
        commands_controls.pack(fill='x', padx=5, pady=5)
        
        # Frame para el botón de inicio/detención de comandos
        command_toggle_frame = ttk.Frame(commands_controls)
        command_toggle_frame.pack(fill='x', pady=5)
        
        # Botón de inicio/detención de comandos con estilo moderno
        self.command_toggle_button = tk.Button(command_toggle_frame, text="Start Commands", 
                                            command=self.toggle_commands,
                                            bg=self.colors['primary'], fg='white',
                                            font=('Segoe UI', 10, 'bold'),
                                            relief='flat', padx=10, pady=5)
        self.command_toggle_button.pack(side='left', padx=5)
        
        # Frame para las expresiones de comando
        command_exprs_frame = ttk.LabelFrame(scrollable_frame, text="Command Expressions", padding=5)
        command_exprs_frame.pack(fill='x', padx=5, pady=5)
        
        # Etiqueta de ayuda
        help_text = """
        Available variables:
        - t: Time (seconds since start)
        - X1-X4: Analog inputs (AN1-AN4)
        - X5: Digital input
        - sin, cos, tan: Trigonometric functions
        - sqrt, abs: Mathematical functions
        - pi: Mathematical constant
        
        Output ranges:
        - Y1-Y4: -5V to 5V
        - Y5: 0 to 255 (digital)
        """
        help_label = ttk.Label(command_exprs_frame, text=help_text, justify='left')
        help_label.pack(fill='x', padx=5, pady=5)
        
        # Crear entradas para cada salida
        for i, output in enumerate(['Y1', 'Y2', 'Y3', 'Y4', 'Y5']):
            expr_frame = ttk.Frame(command_exprs_frame)
            expr_frame.pack(fill='x', pady=2)
            
            ttk.Label(expr_frame, text=f"{output} =").pack(side='left', padx=5)
            entry = ttk.Entry(expr_frame, textvariable=self.command_expressions[output], width=30)
            entry.pack(side='left', padx=5)
            
            apply_btn = tk.Button(expr_frame, text="Apply", 
                                command=lambda e=output: self.apply_command(e),
                                bg=self.colors['primary'], fg='white',
                                font=('Segoe UI', 10, 'bold'),
                                relief='flat', padx=10, pady=2)
            apply_btn.pack(side='left', padx=5)
            
            # Etiqueta para mostrar el rango de salida
            range_text = "(-5V to 5V)" if i < 4 else "(0 to 255)"
            ttk.Label(expr_frame, text=range_text, foreground=self.colors['text_secondary']).pack(side='left', padx=5)
        
    def write_digital(self):
        try:
            value = self.digital_value.get()
            if 0 <= value <= 255:
                if self.daq.write_digital(value, 6):
                    self.update_digital_display(value)
                    messagebox.showinfo("Success", f"Wrote value {value} to byte 6")
                else:
                    messagebox.showerror("Error", "Failed to write value")
            else:
                messagebox.showerror("Error", "Value must be between 0 and 255")
        except ValueError:
            messagebox.showerror("Error", "Please enter a valid number")
            
    def write_analog(self):
        """Write analog value to specified channel"""
        try:
            channel = self.analog_channel.get()
            value = self.analog_value.get()
            if -5 <= value <= 5:
                if self.daq.write_analog(channel, value):
                    messagebox.showinfo("Success", f"Wrote value {value}V to AN{channel}")
                else:
                    messagebox.showerror("Error", "Failed to write value")
            else:
                messagebox.showerror("Error", "Value must be between -5 and 5")
        except ValueError:
            messagebox.showerror("Error", "Please enter valid numbers")
            
    def start_waveform(self):
        """Start waveform generation"""
        try:
            channel = self.analog_channel.get()
            frequency = self.frequency.get()
            amplitude = self.amplitude.get()
            offset = self.offset.get()
            
            if not (0 < frequency <= 100):
                messagebox.showerror("Error", "Frequency must be between 0 and 100 Hz")
                return
                
            if not (0 < amplitude <= 5):
                messagebox.showerror("Error", "Amplitude must be between 0 and 5 V")
                return
                
            if not (-5 <= offset <= 5):
                messagebox.showerror("Error", "Offset must be between -5 and 5 V")
                return
                
            # Iniciar generación de onda en un hilo separado
            self.wave_thread = threading.Thread(
                target=self.daq.write_analog_time_function,
                args=(channel, self.wave_type.get(), frequency, amplitude, offset)
            )
            self.wave_thread.daemon = True
            self.wave_thread.start()
            
            # Actualizar visualización
            self.update_waveform_display()
            
            # Cambiar el botón para detener la onda
            self.wave_button.config(text="Stop Waveform", command=self.stop_waveform)
            
        except ValueError:
            messagebox.showerror("Error", "Please enter valid numbers")
            
    def stop_waveform(self):
        """Stop waveform generation"""
        if hasattr(self, 'wave_thread') and self.wave_thread and self.wave_thread.is_alive():
            self.daq.stop_waveform()  # Usar el nuevo método de detención
            self.wave_button.config(text="Start Waveform", command=self.start_waveform)
            
    def toggle_measurement(self):
        """Toggle measurement state"""
        if not self.measurement_started:
            # Iniciar medición
            self.measurement_started = True
            self.interrupt_button.config(text="Stop Measurement")
            self.start_time = time.perf_counter()  # Usar perf_counter para mayor precisión
            
            # Reiniciar buffer de datos
            self.data_buffer = {
                'time': [],
                'digital_value': [],
                'analog_value': [],
                'Y1': [],
                'Y2': [],
                'Y3': [],
                'Y4': [],
                'Y5': []
            }
            
            # Reiniciar variable t en input_values
            self.input_values['t'] = 0
            
            # Reiniciar last_update_time
            self.last_update_time = 0
            
            self.start_continuous_read()
        else:
            # Detener medición
            self.measurement_started = False
            self.interrupt_button.config(text="Start Measurement")
            self.stop_continuous_read()
            
    def on_tab_changed(self, event):
        """Handle tab change event"""
        # Ya no detenemos la medición al cambiar de pestaña
        pass
            
    def start_continuous_read(self):
        """Start continuous reading of digital and analog values"""
        if not self.reading_active:
            self.reading_active = True
            self.update_gui()  # Iniciar actualización de GUI
            
    def update_gui(self):
        """Update GUI elements"""
        try:
            # Leer datos del DAQ
            self.daq.read_data()
            self.daq.write_data()
            
            # Procesar datos de la cola
            while not self.daq.data_queue.empty():
                try:
                    channel, value = self.daq.data_queue.get_nowait()
                    if channel.startswith('AN'):
                        self._process_analog_data(channel, value)
                    else:
                        self._process_digital_data(value)
                except queue.Empty:
                    break
                    
            # Actualizar buffer de datos si la medición está activa
            if self.measurement_started:
                self.update_data_buffer()
                
                # Actualizar eje de tiempo en las gráficas
                current_time = time.perf_counter() - self.start_time
                self.time_data = np.roll(self.time_data, -1)
                self.time_data[-1] = current_time
                self.graphs_window.update_time(self.time_data)
                
        except Exception as e:
            print(f"Error updating GUI: {e}")
            
        # Programar siguiente actualización
        if self.reading_active:
            self.root.after(10, self.update_gui)  # Actualizar cada 10ms

    def stop_continuous_read(self):
        """Stop continuous reading"""
        self.reading_active = False

    def _process_analog_data(self, channel, value):
        """Process analog data from the DAQ"""
        try:
            channel_num = int(channel[2:])  # Convertir 'AN1' a 1
            
            # Guardar valor en input_values
            self.input_values[f'X{channel_num}'] = value
            
            # Si es el canal seleccionado, actualizar solo la visualización
            if channel_num == self.analog_channel.get():
                # Actualizar la etiqueta de visualización con formato de 3 decimales
                self.analog_value_display.config(text=f"{value:.3f} V")
                
            # Actualizar gráfica
            analog_values = [self.input_values.get(f'X{i}', 0) for i in range(1, 5)]
            self.graphs_window.update_analog_data(analog_values)
                
        except Exception as e:
            print(f"Error processing analog data: {e}")

    def _process_digital_data(self, value):
        """Process digital data"""
        try:
            # Actualizar etiqueta de valor digital
            self.digital_value_label.config(text=f"Value: {value} (0x{value:02X})")
            
            # Actualizar visualización de bits
            self.update_digital_display(value)
            
            # Guardar valor en input_values
            self.input_values['X5'] = value
            
            # Actualizar gráficas
            self.graphs_window.update_digital_data(value)
            self.graphs_window.update_bits_data(value)
            
        except Exception as e:
            print(f"Error processing digital data: {e}")

    def start_waveform_display(self):
        self.wave_canvas.delete("all")
        width = self.wave_canvas.winfo_width()
        height = self.wave_canvas.winfo_height()
        
        # Dibujar ejes
        self.wave_canvas.create_line(0, height/2, width, height/2, fill="gray")  # Eje X
        self.wave_canvas.create_line(0, 0, 0, height, fill="gray")  # Eje Y
        
        # Dibujar onda
        points = []
        for x in range(width):
            t = x / width
            if self.wave_type.get() == "sine":
                y = math.sin(2 * math.pi * self.frequency.get() * t)
            elif self.wave_type.get() == "square":
                y = 1 if math.sin(2 * math.pi * self.frequency.get() * t) >= 0 else -1
            elif self.wave_type.get() == "triangle":
                y = 2 * abs(2 * (self.frequency.get() * t - math.floor(self.frequency.get() * t + 0.5))) - 1
            else:  # sawtooth
                y = 2 * (self.frequency.get() * t - math.floor(self.frequency.get() * t)) - 1
                
            # Escalar y desplazar
            y = y * self.amplitude.get() + self.offset.get()
            y = height/2 - (y / 5) * (height/2)  # Mapear -5V a 5V a la altura del canvas
            
            points.append((x, y))
            
        # Dibujar línea
        if len(points) > 1:
            self.wave_canvas.create_line(points, fill="blue")
            
    def reset_digital(self):
        """Reset digital output to 0"""
        self.digital_value.set(0)
        self.write_digital()
        
    def on_closing(self):
        """Handle window closing"""
        try:
            if hasattr(self, 'reading_active') and self.reading_active:
                self.reading_active = False
            if hasattr(self, 'commands_active') and self.commands_active:
                self.commands_active = False
            if hasattr(self, 'daq'):
                self.daq.close()
            if hasattr(self, 'graphs_window'):
                self.graphs_window.close()
            if hasattr(self, 'root') and self.root:
                self.root.destroy()
        except Exception as e:
            print(f"Error during closing: {e}")
            if hasattr(self, 'root') and self.root:
                self.root.destroy()

    def update_digital_display(self, value):
        # Actualizar etiquetas de bits con colores modernos
        for i in range(8):
            bit = (value >> i) & 1
            self.bit_labels[i].config(
                text=f"Bit {i}: {bit}",
                foreground=self.colors['success'] if bit else self.colors['text_secondary']
            )
            
        # Actualizar visualización en canvas con colores modernos
        self.binary_canvas.delete("all")
        width = self.binary_canvas.winfo_width()
        height = self.binary_canvas.winfo_height()
        
        # Dibujar bits con colores modernos
        for i in range(8):
            bit = (value >> i) & 1
            x = width - (i + 1) * (width / 8)
            color = self.colors['success'] if bit else self.colors['error']
            self.binary_canvas.create_rectangle(x, 0, x + width/8, height, 
                                             fill=color, outline=self.colors['primary'])
            
    def update_waveform_display(self):
        self.wave_canvas.delete("all")
        width = self.wave_canvas.winfo_width()
        height = self.wave_canvas.winfo_height()
        
        # Dibujar ejes
        self.wave_canvas.create_line(0, height/2, width, height/2, fill="gray")  # Eje X
        self.wave_canvas.create_line(0, 0, 0, height, fill="gray")  # Eje Y
        
        # Dibujar onda
        points = []
        for x in range(width):
            t = x / width
            if self.wave_type.get() == "sine":
                y = math.sin(2 * math.pi * self.frequency.get() * t)
            elif self.wave_type.get() == "square":
                y = 1 if math.sin(2 * math.pi * self.frequency.get() * t) >= 0 else -1
            elif self.wave_type.get() == "triangle":
                y = 2 * abs(2 * (self.frequency.get() * t - math.floor(self.frequency.get() * t + 0.5))) - 1
            else:  # sawtooth
                y = 2 * (self.frequency.get() * t - math.floor(self.frequency.get() * t)) - 1
                
            # Escalar y desplazar
            y = y * self.amplitude.get() + self.offset.get()
            y = height/2 - (y / 5) * (height/2)  # Mapear -5V a 5V a la altura del canvas
            
            points.append((x, y))
            
        # Dibujar línea
        if len(points) > 1:
            self.wave_canvas.create_line(points, fill="blue")
            
    def update_digital_graph(self, value):
        """Update the digital value graph using pyqtgraph"""
        try:
            current_time = time.time() - self.start_time if self.start_time else 0
            self.digital_plot.update_data(value, current_time)
            self.bits_plot.update_data(value, current_time)
        except Exception as e:
            print(f"Error updating digital graph: {e}")

    def on_analog_channel_change(self, event):
        """Handle analog channel change"""
        try:
            # Obtener el canal seleccionado
            channel = self.analog_channel.get()
            
            # Actualizar solo la visualización del valor leído
            if f'X{channel}' in self.input_values:
                value = self.input_values[f'X{channel}']
                self.analog_value_display.config(text=f"{value:.3f} V")
            else:
                self.analog_value_display.config(text="0.000 V")
                
        except Exception as e:
            print(f"Error in analog channel change: {e}")

    def update_analog_graph(self, value):
        """Update the analog value graph using pyqtgraph"""
        try:
            # Usar el tiempo actual para la gráfica
            current_time = time.perf_counter()
            self.analog_plot.update_data(value, current_time)
        except Exception as e:
            print(f"Error updating analog graph: {e}")

    def update_time_axis(self):
        """Update time axis for both graphs"""
        if self.measurement_started and self.start_time is not None:
            current_time = time.time() - self.start_time
            
            # Actualizar time_data solo cuando sea necesario
            if current_time > self.last_update_time + self.read_sampling_period:
                # Desplazar el array de tiempo
                self.time_data = np.roll(self.time_data, -1)
                self.time_data[-1] = current_time
                self.last_update_time = current_time
                
                # Actualizar ejes X
                self.digital_ax.set_xlim(self.time_data[0], self.time_data[-1])
                self.analog_ax.set_xlim(self.time_data[0], self.time_data[-1])
                self.bits_ax.set_xlim(self.time_data[0], self.time_data[-1])
                
                # Actualizar datos de tiempo
                self.digital_line.set_xdata(self.time_data)
                self.analog_line.set_xdata(self.time_data)
                for line in self.bits_lines:
                    line.set_xdata(self.time_data)
                
                # Forzar actualización de los canvas
                self.digital_canvas.draw()
                self.analog_canvas.draw()
                self.bits_canvas.draw()

    def toggle_commands(self):
        """Toggle command execution state"""
        if not self.commands_active:
            # Iniciar comandos
            self.commands_active = True
            self.command_toggle_button.config(text="Stop Commands")
            self.command_start_time = time.perf_counter()
            self.start_command_execution()
        else:
            # Detener comandos
            self.commands_active = False
            self.command_toggle_button.config(text="Start Commands")
            self.stop_command_execution()
            
    def start_command_execution(self):
        """Start continuous command execution"""
        if not self.commands_active:
            return
            
        try:
            # Actualizar variable t
            current_time = time.perf_counter() - self.command_start_time
            self.input_values['t'] = current_time
            
            # Ejecutar cada comando
            for output in ['Y1', 'Y2', 'Y3', 'Y4', 'Y5']:
                try:
                    expr = self.command_expressions[output].get()
                    result = eval(expr, {"__builtins__": {}}, self.input_values)
                    
                    # Limitar el rango según el tipo de salida
                    if output in ['Y1', 'Y2', 'Y3', 'Y4']:
                        result = max(min(result, 5), -5)  # Limitar entre -5V y 5V
                    else:  # Y5
                        result = max(min(int(result), 255), 0)  # Limitar entre 0 y 255
                    
                    # Aplicar resultado
                    if output in ['Y1', 'Y2', 'Y3', 'Y4']:
                        self.daq.write_analog(int(output[1]), result)
                    else:  # Y5
                        self.daq.write_digital(int(result), 6)
                                
                except Exception as e:
                    print(f"Error executing command for {output}: {e}")
                    
        except Exception as e:
            print(f"Error in command execution: {e}")
            
        # Programar siguiente ejecución
        if self.commands_active:
            self.root.after(10, self.start_command_execution)  # Actualizar cada 10ms
            
    def stop_command_execution(self):
        """Stop continuous command execution"""
        self.commands_active = False
        
    def apply_command(self, output):
        """Apply a single command"""
        try:
            expr = self.command_expressions[output].get()
            result = eval(expr, {"__builtins__": {}}, self.input_values)
            
            # Limitar el rango según el tipo de salida
            if output in ['Y1', 'Y2', 'Y3', 'Y4']:
                result = max(min(result, 5), -5)  # Limitar entre -5V y 5V
            else:  # Y5
                result = max(min(int(result), 255), 0)  # Limitar entre 0 y 255
            
            # Aplicar resultado
            if output in ['Y1', 'Y2', 'Y3', 'Y4']:
                self.daq.write_analog(int(output[1]), result)
            else:  # Y5
                self.daq.write_digital(int(result), 6)
                
            messagebox.showinfo("Success", f"Applied command for {output}")
            
        except Exception as e:
            messagebox.showerror("Error", f"Invalid command for {output}: {str(e)}")

    def update_data_buffer(self):
        """Update data buffer with current values"""
        if self.measurement_started:
            try:
                current_time = time.perf_counter() - self.start_time  # Usar perf_counter para mayor precisión
                
                # Actualizar buffer con datos de inputs
                self.data_buffer['time'].append(current_time)
                self.data_buffer['digital_value'].append(self.input_values.get('X5', 0))
                
                # Guardar valores históricos de entradas analógicas
                for i in range(1, 5):
                    self.data_buffer[f'X{i}'] = self.data_buffer.get(f'X{i}', [])
                    self.data_buffer[f'X{i}'].append(self.input_values.get(f'X{i}', 0))
                
                # Actualizar variable t en input_values
                self.input_values['t'] = current_time
                
                # Calcular y guardar valores de comandos
                output_values = []
                for output in ['Y1', 'Y2', 'Y3', 'Y4']:
                    try:
                        expr = self.command_expressions[output].get()
                        result = eval(expr, {"__builtins__": {}}, self.input_values)
                        result = max(min(result, 5), -5)  # Limitar entre -5V y 5V
                        self.data_buffer[output].append(result)
                        output_values.append(result)
                    except Exception as e:
                        print(f"Error calculating {output}: {e}")
                        self.data_buffer[output].append(0)
                        output_values.append(0)
                
                # Calcular y guardar valor digital Y5
                try:
                    expr = self.command_expressions['Y5'].get()
                    result = eval(expr, {"__builtins__": {}}, self.input_values)
                    result = max(min(int(result), 255), 0)  # Limitar entre 0 y 255
                    self.data_buffer['Y5'].append(result)
                except Exception as e:
                    print(f"Error calculating Y5: {e}")
                    self.data_buffer['Y5'].append(0)
                
                # Actualizar gráficas de salidas
                self.graphs_window.update_output_data(output_values)
                self.graphs_window.update_digital_output_data(self.data_buffer['Y5'][-1])
                
                # Mantener solo los últimos max_points datos
                if len(self.data_buffer['time']) > self.max_points:
                    for key in self.data_buffer:
                        self.data_buffer[key] = self.data_buffer[key][-self.max_points:]
                        
            except Exception as e:
                print(f"Error updating data buffer: {e}")

    def save_data_to_excel(self):
        """Save data to Excel file"""
        try:
            if not self.measurement_started:
                messagebox.showwarning("Warning", "No measurement data available. Start a measurement first.")
                return
                
            # Detener la medición automáticamente
            self.measurement_started = False
            self.interrupt_button.config(text="Start Measurement")
            self.stop_continuous_read()
                
            # Verificar si hay datos suficientes
            if not self.data_buffer['time']:
                messagebox.showwarning("Warning", "No data available to save. Start a measurement and collect some data first.")
                return
                
            # Solicitar ubicación para guardar el archivo
            filename = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx")],
                title="Save Data As"
            )
            
            if not filename:
                return
                
            # Crear archivo Excel usando openpyxl
            wb = Workbook()
            
            # Hoja de datos
            ws_data = wb.active
            ws_data.title = "Data"
            
            # Escribir encabezados
            headers = [
                'Time (s)',
                'Digital Input (X5)',
                'Analog Input AN1 (X1)',
                'Analog Input AN2 (X2)',
                'Analog Input AN3 (X3)',
                'Analog Input AN4 (X4)',
                'Analog Output Y1',
                'Analog Output Y2',
                'Analog Output Y3',
                'Analog Output Y4',
                'Digital Output Y5'
            ]
            for col, header in enumerate(headers, 1):
                ws_data.cell(row=1, column=col, value=header)
            
            # Escribir datos
            for row_idx, i in enumerate(range(len(self.data_buffer['time'])), 2):
                ws_data.cell(row=row_idx, column=1, value=self.data_buffer['time'][i])
                ws_data.cell(row=row_idx, column=2, value=self.data_buffer['digital_value'][i])
                ws_data.cell(row=row_idx, column=3, value=self.data_buffer.get('X1', [0] * len(self.data_buffer['time']))[i])
                ws_data.cell(row=row_idx, column=4, value=self.data_buffer.get('X2', [0] * len(self.data_buffer['time']))[i])
                ws_data.cell(row=row_idx, column=5, value=self.data_buffer.get('X3', [0] * len(self.data_buffer['time']))[i])
                ws_data.cell(row=row_idx, column=6, value=self.data_buffer.get('X4', [0] * len(self.data_buffer['time']))[i])
                ws_data.cell(row=row_idx, column=7, value=self.data_buffer['Y1'][i])
                ws_data.cell(row=row_idx, column=8, value=self.data_buffer['Y2'][i])
                ws_data.cell(row=row_idx, column=9, value=self.data_buffer['Y3'][i])
                ws_data.cell(row=row_idx, column=10, value=self.data_buffer['Y4'][i])
                ws_data.cell(row=row_idx, column=11, value=self.data_buffer['Y5'][i])
            
            # Hoja de metadatos
            ws_meta = wb.create_sheet("Metadata")
            ws_meta.cell(row=1, column=1, value="Parameter")
            ws_meta.cell(row=1, column=2, value="Value")
            ws_meta.cell(row=2, column=1, value="Measurement Start Time")
            ws_meta.cell(row=2, column=2, value=datetime.fromtimestamp(self.start_time).strftime('%Y-%m-%d %H:%M:%S'))
            ws_meta.cell(row=3, column=1, value="Measurement Duration (s)")
            ws_meta.cell(row=3, column=2, value=self.data_buffer['time'][-1])
            ws_meta.cell(row=4, column=1, value="Sampling Period (s)")
            ws_meta.cell(row=4, column=2, value=self.read_sampling_period)
            ws_meta.cell(row=5, column=1, value="Total Samples")
            ws_meta.cell(row=5, column=2, value=len(self.data_buffer['time']))
            
            # Ajustar ancho de columnas
            for ws in [ws_data, ws_meta]:
                for column in ws.columns:
                    max_length = 0
                    column = list(column)
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = (max_length + 2)
                    ws.column_dimensions[column[0].column_letter].width = adjusted_width
            
            # Guardar archivo
            wb.save(filename)
            
            messagebox.showinfo("Success", "Data saved successfully!")
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to save data: {str(e)}")
            print(f"Error details: {str(e)}")  # Para debugging

    def on_window_move(self, event):
        """Synchronize the position of the PyQt window with the main window"""
        if event.widget == self.root:
            x = self.root.winfo_x()
            y = self.root.winfo_y()
            width = self.root.winfo_width()
            self.graphs_window.move(x + width + 20, y)  # 20 pixels de separación

class GraphsWindow(QMainWindow):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("DAQ Graphs")
        self.setGeometry(100, 100, 1200, 800)
        
        # Widget central
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        layout = QVBoxLayout(central_widget)
        
        # Controles
        controls_layout = QHBoxLayout()
        
        # Frame para controles de entradas analógicas
        analog_controls = QVBoxLayout()
        analog_label = QLabel("Analog Inputs:")
        analog_controls.addWidget(analog_label)
        
        # Checkboxes para canales analógicos de entrada
        analog_checkbox_layout = QHBoxLayout()
        self.analog_channel_checks = []
        colors = ['r', 'g', 'b', 'y']
        for i in range(4):
            checkbox = QCheckBox(f"AN{i+1}")
            checkbox.setChecked(True)
            checkbox.setStyleSheet(f"color: {colors[i]};")
            checkbox.stateChanged.connect(lambda state, idx=i: self.toggle_analog_channel(idx, state))
            analog_checkbox_layout.addWidget(checkbox)
            self.analog_channel_checks.append(checkbox)
        analog_controls.addLayout(analog_checkbox_layout)
        
        # Frame para controles de salidas analógicas
        output_controls = QVBoxLayout()
        output_label = QLabel("Analog Outputs:")
        output_controls.addWidget(output_label)
        
        # Checkboxes para salidas analógicas
        output_checkbox_layout = QHBoxLayout()
        self.output_channel_checks = []
        output_colors = ['c', 'm', 'w', 'g']  # Colores diferentes para las salidas
        for i in range(4):
            checkbox = QCheckBox(f"Y{i+1}")
            checkbox.setChecked(True)
            checkbox.setStyleSheet(f"color: {output_colors[i]};")
            checkbox.stateChanged.connect(lambda state, idx=i: self.toggle_output_channel(idx, state))
            output_checkbox_layout.addWidget(checkbox)
            self.output_channel_checks.append(checkbox)
        output_controls.addLayout(output_checkbox_layout)
        
        # Frame para otros controles
        other_controls = QVBoxLayout()
        other_label = QLabel("Other Channels:")
        other_controls.addWidget(other_label)
        
        # Checkboxes para digital y bits
        self.digital_check = QCheckBox("Digital Input")
        self.digital_check.setChecked(True)
        self.digital_check.stateChanged.connect(self.toggle_digital_graph)
        other_controls.addWidget(self.digital_check)
        
        self.bits_check = QCheckBox("Digital Bits")
        self.bits_check.setChecked(True)
        self.bits_check.stateChanged.connect(self.toggle_bits_graph)
        other_controls.addWidget(self.bits_check)
        
        # Checkbox para salida digital
        self.digital_output_check = QCheckBox("Digital Output (Y5)")
        self.digital_output_check.setChecked(True)
        self.digital_output_check.stateChanged.connect(self.toggle_digital_output_graph)
        other_controls.addWidget(self.digital_output_check)
        
        # Agregar frames de controles al layout principal
        controls_layout.addLayout(analog_controls)
        controls_layout.addLayout(output_controls)
        controls_layout.addLayout(other_controls)
        layout.addLayout(controls_layout)
        
        # Gráfica de entradas analógicas
        self.analog_widget = pg.PlotWidget()
        self.analog_widget.setTitle("Analog Inputs")
        self.analog_widget.setLabel('left', 'Voltage (V)')
        self.analog_widget.setLabel('bottom', 'Time (s)')
        self.analog_widget.addLegend()
        self.analog_curves = []
        for i in range(4):
            curve = self.analog_widget.plot(pen=colors[i], name=f'AN{i+1}')
            self.analog_curves.append(curve)
        layout.addWidget(self.analog_widget)
        
        # Gráfica de salidas analógicas
        self.output_widget = pg.PlotWidget()
        self.output_widget.setTitle("Analog Outputs")
        self.output_widget.setLabel('left', 'Voltage (V)')
        self.output_widget.setLabel('bottom', 'Time (s)')
        self.output_widget.addLegend()
        self.output_curves = []
        for i in range(4):
            curve = self.output_widget.plot(pen=output_colors[i], name=f'Y{i+1}')
            self.output_curves.append(curve)
        layout.addWidget(self.output_widget)
        
        # Gráfica de valor digital
        self.digital_widget = pg.PlotWidget()
        self.digital_widget.setTitle("Digital Input")
        self.digital_widget.setLabel('left', 'Value')
        self.digital_widget.setLabel('bottom', 'Time (s)')
        self.digital_curve = self.digital_widget.plot(pen='w')
        layout.addWidget(self.digital_widget)
        
        # Gráfica de salida digital
        self.digital_output_widget = pg.PlotWidget()
        self.digital_output_widget.setTitle("Digital Output (Y5)")
        self.digital_output_widget.setLabel('left', 'Value')
        self.digital_output_widget.setLabel('bottom', 'Time (s)')
        self.digital_output_curve = self.digital_output_widget.plot(pen='y')
        layout.addWidget(self.digital_output_widget)
        
        # Gráfica de bits digitales (analizador lógico)
        self.bits_widget = pg.PlotWidget()
        self.bits_widget.setTitle("Digital Bits")
        self.bits_widget.setLabel('left', 'Bit Level')
        self.bits_widget.setLabel('bottom', 'Time (s)')
        self.bits_widget.addLegend()
        
        # Configurar ejes Y para mostrar bits en niveles
        self.bits_widget.setYRange(-0.5, 7.5)  # 8 bits, de 0 a 7
        self.bits_widget.getAxis('left').setTicks([[(i, f'Bit {i}') for i in range(8)]])
        
        # Crear una sola curva para cada bit
        self.bits_curves = []
        for i in range(8):
            curve = self.bits_widget.plot(pen=colors[i % len(colors)], name=f'Bit {i}')
            self.bits_curves.append(curve)
        layout.addWidget(self.bits_widget)
        
        # Datos para las gráficas
        self.time_data = np.zeros(1000)
        self.analog_data = [np.zeros(1000) for _ in range(4)]
        self.output_data = [np.zeros(1000) for _ in range(4)]
        self.digital_data = np.zeros(1000)
        self.digital_output_data = np.zeros(1000)
        self.bits_data = [np.zeros(1000) for _ in range(8)]
        
    def toggle_analog_channel(self, channel_idx, state):
        """Toggle visibility of a specific analog input channel"""
        if state == Qt.Checked:
            self.analog_curves[channel_idx].setData(self.time_data, self.analog_data[channel_idx])
        else:
            self.analog_curves[channel_idx].setData([], [])
            
    def toggle_output_channel(self, channel_idx, state):
        """Toggle visibility of a specific analog output channel"""
        if state == Qt.Checked:
            self.output_curves[channel_idx].setData(self.time_data, self.output_data[channel_idx])
        else:
            self.output_curves[channel_idx].setData([], [])
            
    def toggle_digital_graph(self, state):
        self.digital_widget.setVisible(state == Qt.Checked)
        
    def toggle_digital_output_graph(self, state):
        self.digital_output_widget.setVisible(state == Qt.Checked)
        
    def toggle_bits_graph(self, state):
        self.bits_widget.setVisible(state == Qt.Checked)
        
    def update_analog_data(self, values):
        """Update analog input graph data"""
        self.analog_data = [np.roll(data, -1) for data in self.analog_data]
        for i, value in enumerate(values):
            self.analog_data[i][-1] = value
            if self.analog_channel_checks[i].isChecked():
                self.analog_curves[i].setData(self.time_data, self.analog_data[i])
            
    def update_output_data(self, values):
        """Update analog output graph data"""
        self.output_data = [np.roll(data, -1) for data in self.output_data]
        for i, value in enumerate(values):
            self.output_data[i][-1] = value
            if self.output_channel_checks[i].isChecked():
                self.output_curves[i].setData(self.time_data, self.output_data[i])
            
    def update_digital_data(self, value):
        """Update digital input graph data"""
        self.digital_data = np.roll(self.digital_data, -1)
        self.digital_data[-1] = value
        self.digital_curve.setData(self.time_data, self.digital_data)
        
    def update_digital_output_data(self, value):
        """Update digital output graph data"""
        self.digital_output_data = np.roll(self.digital_output_data, -1)
        self.digital_output_data[-1] = value
        self.digital_output_curve.setData(self.time_data, self.digital_output_data)
        
    def update_bits_data(self, value):
        """Update digital bits graph data"""
        for i in range(8):
            bit = (value >> i) & 1
            self.bits_data[i] = np.roll(self.bits_data[i], -1)
            self.bits_data[i][-1] = bit
            # Mostrar el bit en su nivel correspondiente (0-7)
            self.bits_curves[i].setData(self.time_data, self.bits_data[i] + i)
            
    def update_time(self, time_data):
        """Update time axis for all graphs"""
        self.time_data = time_data
        self.analog_widget.setXRange(time_data[0], time_data[-1])
        self.output_widget.setXRange(time_data[0], time_data[-1])
        self.digital_widget.setXRange(time_data[0], time_data[-1])
        self.digital_output_widget.setXRange(time_data[0], time_data[-1])
        self.bits_widget.setXRange(time_data[0], time_data[-1])

def main():
    root = tk.Tk()
    app = DAQGUI(root)
    root.mainloop()

if __name__ == "__main__":
    main() 